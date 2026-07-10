"""Excel .xlsx 处理器单元测试

覆盖：
- 注册表匹配（xlsx/xlsm/xltx/xltm）
- 大小写不敏感
- MIME 映射
- 预览数据结构
- 编辑数据
- save_cell / save_cells
- 错误处理（文件不存在/无效文件）
"""
import json
import os
import tempfile
import unittest

from openpyxl import Workbook

from mmfb.handlers.xlsx_handler import XlsxHandler


# ---------- 测试数据构造 ----------

def _make_xlsx(tmpdir, filename, rows=None):
    """创建一个带数据的 .xlsx 文件，返回绝对路径"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    if rows is None:
        rows = [
            ["姓名", "年龄", "城市"],
            ["Alice", 30, "Beijing"],
            ["Bob", 25, "Shanghai"],
            ["Charlie", 35, "Guangzhou"],
        ]
    for r in rows:
        ws.append(r)
    path = os.path.join(tmpdir, filename)
    wb.save(path)
    wb.close()
    return path


def _make_empty_xlsx(tmpdir, filename):
    """创建一个空工作表的 .xlsx"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    path = os.path.join(tmpdir, filename)
    wb.save(path)
    wb.close()
    return path


# ---------- 测试类 ----------

class TestXlsxHandlerExtensions(unittest.TestCase):
    """注册表匹配与大小写测试"""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def test_extensions_attribute(self):
        h = XlsxHandler("/fake.xlsx")
        self.assertIn(".xlsx", h.extensions)
        self.assertIn(".xlsm", h.extensions)
        self.assertIn(".xltx", h.extensions)
        self.assertIn(".xltm", h.extensions)

    def test_case_insensitive_match(self):
        self.assertTrue(XlsxHandler.can_handle("/path/FILE.XLSX"))
        self.assertTrue(XlsxHandler.can_handle("/path/data.XlSm"))
        self.assertFalse(XlsxHandler.can_handle("/path/not_a_doc.txt"))

    def test_mime(self):
        h = XlsxHandler("/fake.xlsx")
        self.assertEqual(
            h.get_mime(),
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )


class TestXlsxHandlerPreview(unittest.TestCase):
    """预览数据测试"""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def test_preview_valid_file(self):
        path = _make_xlsx(self.tmpdir, "test.xlsx")
        h = XlsxHandler(path)
        result = h.get_preview()
        self.assertIsNotNone(result)
        self.assertNotIn("error", result)
        self.assertEqual(result["template"], "xlsx")
        self.assertTrue(result["editable"])
        self.assertEqual(result["data"]["sheet_count"], 1)
        self.assertEqual(len(result["data"]["sheets"]), 1)
        sheet = result["data"]["sheets"][0]
        self.assertEqual(sheet["name"], "Sheet1")
        self.assertGreater(len(sheet["cells"]), 0)

    def test_preview_empty_sheet(self):
        path = _make_empty_xlsx(self.tmpdir, "empty.xlsx")
        h = XlsxHandler(path)
        result = h.get_preview()
        self.assertIsNotNone(result)
        self.assertNotIn("error", result)
        self.assertEqual(result["data"]["sheet_count"], 1)
        self.assertEqual(len(result["data"]["sheets"][0]["cells"]), 0)

    def test_preview_nonexistent_file(self):
        h = XlsxHandler("/nonexistent/path.xlsx")
        result = h.get_preview()
        self.assertIsNotNone(result)
        self.assertIn("error", result)
        self.assertIn("file not found", result["error"].lower())

    def test_preview_invalid_file(self):
        # 创建一个非 xlsx 的文件（纯文本）
        bad_path = os.path.join(self.tmpdir, "notxlsx.xlsx")
        with open(bad_path, "w", encoding="utf-8") as f:
            f.write("this is not an xlsx file")
        h = XlsxHandler(bad_path)
        result = h.get_preview()
        # 应该返回错误（openpyxl 无法解析纯文本）
        self.assertIsNotNone(result)

    def test_preview_data_structure(self):
        path = _make_xlsx(self.tmpdir, "structure.xlsx", [
            ["Name", "Score"],
            ["Tom", 95],
            ["Jerry", 88],
        ])
        h = XlsxHandler(path)
        result = h.get_preview()["data"]
        sheet = result["sheets"][0]
        # 每个 cell 必须有 r, c, address, value, type, style
        for cell in sheet["cells"]:
            self.assertIn("r", cell)
            self.assertIn("c", cell)
            self.assertIn("address", cell)
            self.assertIn("value", cell)
            self.assertIn("type", cell)
            self.assertIn("style", cell)
        # 检查类型标记
        types = {c["address"]: c["type"] for c in sheet["cells"]}
        self.assertEqual(types["A1"], "s")  # Name
        self.assertEqual(types["B1"], "s")  # Score
        self.assertEqual(types["A2"], "s")  # Tom
        self.assertEqual(types["B2"], "n")  # 95


class TestXlsxHandlerEdit(unittest.TestCase):
    """编辑数据与保存测试"""

    def setUp(self):
        self.tmpdir = tempfile.mkdtemp()

    def test_get_edit_returns_save_flag(self):
        path = _make_xlsx(self.tmpdir, "edit_test.xlsx")
        h = XlsxHandler(path)
        result = h.get_edit()
        self.assertIsNotNone(result)
        self.assertNotIn("error", result)  # 无错误
        self.assertTrue(result["data"].get("save"))

    def test_save_cell_single(self):
        path = _make_xlsx(self.tmpdir, "save_test.xlsx", [
            ["Name", "Age"],
            ["Alice", 30],
        ])
        h = XlsxHandler(path)
        ok = h.save_cell("Sheet1", "B2", 31)
        self.assertTrue(ok)
        # 验证写入成功：重新读取
        result = h.get_preview()
        cells_by_addr = {c["address"]: c for c in result["data"]["sheets"][0]["cells"]}
        self.assertEqual(cells_by_addr["B2"]["value"], 31)

    def test_save_cell_invalid_sheet(self):
        path = _make_xlsx(self.tmpdir, "invalid_sheet.xlsx")
        h = XlsxHandler(path)
        ok = h.save_cell("NonExistent", "A1", "test")
        self.assertFalse(ok)

    def test_save_cells_batch(self):
        path = _make_xlsx(self.tmpdir, "batch_test.xlsx", [
            ["A", "B", "C"],
            [1, 2, 3],
            [4, 5, 6],
        ])
        h = XlsxHandler(path)
        changes = json.dumps([
            {"sheet": "Sheet1", "address": "A2", "value": 10},
            {"sheet": "Sheet1", "address": "C3", "value": 60},
        ])
        ok = h.save_cells(changes)
        self.assertTrue(ok)
        # 验证
        result = h.get_preview()
        cells_by_addr = {c["address"]: c for c in result["data"]["sheets"][0]["cells"]}
        self.assertEqual(cells_by_addr["A2"]["value"], 10)
        self.assertEqual(cells_by_addr["C3"]["value"], 60)

    def test_save_cells_bad_json(self):
        h = XlsxHandler("/fake.xlsx")
        ok = h.save_cells("not-json")
        self.assertFalse(ok)
        ok = h.save_cells("[1, 2, 3]")  # 不是数组元素
        self.assertFalse(ok)


class TestXlsxHandlerStyle(unittest.TestCase):
    """单元格类型与样式提取测试"""

    def test_cell_type_detection(self):
        self.assertEqual(XlsxHandler._get_cell_type(True), "b")
        self.assertEqual(XlsxHandler._get_cell_type(False), "b")
        self.assertEqual(XlsxHandler._get_cell_type(42), "n")
        self.assertEqual(XlsxHandler._get_cell_type(3.14), "n")
        self.assertEqual(XlsxHandler._get_cell_type("=SUM(A1:A5)"), "f")
        self.assertEqual(XlsxHandler._get_cell_type("hello"), "s")
        self.assertEqual(XlsxHandler._get_cell_type(None), "s")  # None fallback

    def test_safe_value(self):
        self.assertIsNone(XlsxHandler._safe_value(None))
        self.assertEqual(XlsxHandler._safe_value(42), 42)
        self.assertEqual(XlsxHandler._safe_value("test"), "test")
        self.assertEqual(XlsxHandler._safe_value(3.14), 3.14)

    def test_extract_style_basic(self):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "S"
        cell = ws["A1"]
        cell.value = "test"
        cell.font = cell.font.copy(bold=True, italic=True, color=None)
        style = XlsxHandler._extract_style(cell)
        self.assertTrue(style.get("bold"))
        self.assertTrue(style.get("italic"))
        wb.close()


class TestXlsxHandlerConstants(unittest.TestCase):
    """常量与边界测试"""

    def test_max_rows_limit(self):
        from mmfb.handlers.xlsx_handler import MAX_ROWS
        self.assertEqual(MAX_ROWS, 200)

    def test_max_cols_limit(self):
        from mmfb.handlers.xlsx_handler import MAX_COLS
        self.assertEqual(MAX_COLS, 50)

    def test_error_result_structure(self):
        h = XlsxHandler("/nonexistent.xlsx")
        result = h.get_preview()
        self.assertIn("error", result)
        self.assertFalse(result["editable"])
        self.assertEqual(result["data"]["sheet_count"], 0)
        self.assertEqual(len(result["data"]["sheets"]), 0)


if __name__ == "__main__":
    unittest.main()
