"""Excel .xlsx 格式处理器

职责：
1. 使用 openpyxl 读取 xlsx 文件，提取表页结构/单元格数据/样式
2. 预览：返回 JSON 供前端原生 table 渲染（不依赖 sheetjs，保持轻量）
3. 编辑：接收单元格坐标（sheet/address/value）并写回 xlsx
4. 大文件限制：最多取 MAX_ROWS x MAX_COLS，超出部分截断并标记

数据格式（预览返回 data.sheets[i].cells[r][c]）：
    { address: "A1", value: any, type: "s|n|b|f", style: {bold,italic,color} }
"""
import json
import os
from typing import Any, Dict, List, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from mmfb.core.handler_base import BaseHandler


# 最大提取规模：200 行 x 50 列 = 10000 单元格，保证预览流畅
MAX_ROWS = 200
MAX_COLS = 50

XLSX_EXTENSIONS = [".xlsx", ".xlsm", ".xltx", ".xltm"]


class XlsxHandler(BaseHandler):
    """Excel .xlsx / .xlsm / .xltx / .xltm 处理器

    预览模式：返回 sheets 结构数据，前端用原生 table 渲染
    编辑模式：通过 save_cell 桥接写入单元格
    """

    extensions = XLSX_EXTENSIONS

    def get_preview(self) -> Optional[Dict[str, Any]]:
        """获取 xlsx 预览数据"""
        try:
            if not os.path.isfile(self.path):
                return self._make_error("file not found")

            file_size = os.path.getsize(self.path)
            try:
                wb = load_workbook(self.path, read_only=False, data_only=False)
            except Exception as e:
                return self._make_error(f"open failed: {e}", file_size)

            sheets = []
            for ws_name in wb.sheetnames:
                ws = wb[ws_name]
                cells, max_r, max_c = self._extract_cells(ws)
                sheets.append({
                    "name": ws_name,
                    "title": ws.title,
                    "maxRow": max_r,
                    "maxCol": max_c,
                    "cells": cells,
                })

            wb.close()

            return {
                "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                "template": "xlsx",
                "data": {
                    "file_path": os.path.abspath(self.path),
                    "file_size": file_size,
                    "sheets": sheets,
                    "sheet_count": len(sheets),
                    "editable": True,
                },
                "editable": True,
            }
        except Exception as e:
            return self._make_error(str(e))

    def get_edit(self) -> Optional[Dict[str, Any]]:
        """获取编辑数据（与预览相同，额外标记 save 标志）"""
        preview = self.get_preview()
        if preview is None:
            return None
        if "error" in preview:
            return preview
        preview["data"]["save"] = True
        return preview

    def save_cell(self, sheet_name: str, address: str, value: Any) -> bool:
        """单个单元格写入

        Args:
            sheet_name: 工作表名称
            address: 单元格地址（如 "A1"）
            value: 新值（str | int | float | None）
        """
        try:
            wb = load_workbook(self.path)
            if sheet_name not in wb.sheetnames:
                return False
            ws = wb[sheet_name]
            ws[address] = value
            wb.save(self.path)
            wb.close()
            return True
        except Exception as e:
            print(f"[XlsxHandler] save_cell error: {e}")
            return False

    def save_cells(self, changes_json: str) -> bool:
        """批量写入单元格

        Args:
            changes_json: JSON 字符串 [{ sheet, address, value }, ...]
        """
        try:
            changes = json.loads(changes_json)
            if not isinstance(changes, list):
                return False
            wb = load_workbook(self.path)
            for ch in changes:
                sn = ch.get("sheet", "")
                addr = ch.get("address", "")
                val = ch.get("value")
                if sn in wb.sheetnames and addr:
                    wb[sn][addr] = val
            wb.save(self.path)
            wb.close()
            return True
        except Exception as e:
            print(f"[XlsxHandler] save_cells error: {e}")
            return False

    def _extract_cells(self, ws):
        """从工作表提取单元格数据，限制 MAX_ROWS x MAX_COLS

        Returns:
            (cells_flat, max_row_used, max_col_used)
            cells_flat: [{r, c, address, value, type, style}, ...]
        """
        cells = []
        max_r = 0
        max_c = 0

        # 遍历实际有数据的行（read_only 模式返回生成器）
        for row_idx, row in enumerate(ws.iter_rows(
            max_row=MAX_ROWS, max_col=MAX_COLS
        ), start=1):
            for col_idx, cell in enumerate(row, start=1):
                val = cell.value
                if val is None:
                    continue
                max_r = max(max_r, row_idx)
                max_c = max(max_c, col_idx)

                cell_type = self._get_cell_type(val)
                style = self._extract_style(cell)

                cells.append({
                    "r": row_idx - 1,   # 前端 0-indexed
                    "c": col_idx - 1,
                    "address": f"{get_column_letter(col_idx)}{row_idx}",
                    "value": self._safe_value(val),
                    "type": cell_type,
                    "style": style,
                })

        return cells, max_r, max_c

    @staticmethod
    def _get_cell_type(val: Any) -> str:
        """返回类型标记: s=string, n=number, b=boolean, f=formula, d=date"""
        if isinstance(val, bool):
            return "b"
        if isinstance(val, (int, float)):
            return "n"
        if isinstance(val, str) and val.startswith("="):
            return "f"
        # datetime 类型
        type_name = type(val).__name__
        if "date" in type_name.lower():
            return "d"
        return "s"

    @staticmethod
    def _safe_value(val: Any) -> Any:
        """确保值可 JSON 序列化"""
        if val is None:
            return None
        if isinstance(val, (str, int, float, bool)):
            return val
        # datetime 等转字符串
        return str(val)

    @staticmethod
    def _extract_style(cell) -> Dict[str, Any]:
        """提取单元格关键样式"""
        style: Dict[str, Any] = {}
        try:
            wb = cell.parent.parent
            if cell.font:
                if cell.font.bold:
                    style["bold"] = True
                if cell.font.italic:
                    style["italic"] = True
                if cell.font.color:
                    color_val = _resolve_color(wb, cell.font.color)
                    if color_val:
                        style["color"] = color_val
            if cell.fill and cell.fill.fgColor:
                bg_val = _resolve_color(wb, cell.fill.fgColor)
                if bg_val:
                    style["bgColor"] = bg_val
            if cell.alignment:
                if cell.alignment.horizontal == "center":
                    style["align"] = "center"
                elif cell.alignment.horizontal == "right":
                    style["align"] = "right"
        except Exception:
            pass
        return style

    def _make_error(self, error_msg: str, file_size: int = 0) -> Dict[str, Any]:
        """构造错误返回"""
        return {
            "mime": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "template": "xlsx",
            "data": {
                "file_path": os.path.abspath(self.path),
                "file_size": file_size,
                "sheets": [],
                "sheet_count": 0,
                "editable": False,
            },
            "editable": False,
            "error": error_msg,
        }

def _get_theme_color_from_wb(wb, theme_idx) -> Optional[str]:
    theme = getattr(wb, 'theme', None)
    if not theme:
        return None
    try:
        if isinstance(theme, (str, bytes)):
            import xml.etree.ElementTree as ET
            root = ET.fromstring(theme)
            ns = {'a': 'http://schemas.openxmlformats.org/drawingml/2006/main'}
            clr_scheme = root.find('.//a:clrScheme', ns)
            if clr_scheme is not None:
                color_names = ["lt1", "dk1", "lt2", "dk2", "accent1", "accent2", "accent3", "accent4", "accent5", "accent6", "hlink", "folHlink"]
                if 0 <= theme_idx < len(color_names):
                    name = color_names[theme_idx]
                    color_node = clr_scheme.find(f'./a:{name}', ns)
                    if color_node is not None:
                        srgb = color_node.find('./a:srgbClr', ns)
                        if srgb is not None and 'val' in srgb.attrib:
                            return srgb.attrib['val']
                        sys_clr = color_node.find('./a:sysClr', ns)
                        if sys_clr is not None and 'lastClr' in sys_clr.attrib:
                            return sys_clr.attrib['lastClr']
    except Exception:
        pass
    return None

def _resolve_color(wb, color_obj) -> Optional[str]:
    if color_obj is None:
        return None
    rgb = None
    if color_obj.type == 'rgb' and color_obj.rgb:
        rgb = str(color_obj.rgb)
    elif color_obj.type == 'indexed' and color_obj.indexed is not None:
        try:
            from openpyxl.styles.colors import COLOR_INDEX
            if 0 <= color_obj.indexed < len(COLOR_INDEX):
                rgb = COLOR_INDEX[color_obj.indexed]
        except Exception:
            pass
    elif color_obj.type == 'theme' and color_obj.theme is not None:
        rgb = _get_theme_color_from_wb(wb, color_obj.theme)
        if not rgb:
            DEFAULT_THEME_COLORS = [
                "FFFFFF", "000000", "EEECE1", "1F497D",
                "4F81BD", "C0504D", "9BBB59", "8064A2",
                "4BACC6", "F79646", "0000FF", "800080"
            ]
            if 0 <= color_obj.theme < len(DEFAULT_THEME_COLORS):
                rgb = DEFAULT_THEME_COLORS[color_obj.theme]
    if not rgb:
        return None
    rgb = rgb.replace("#", "").replace(":", "").upper()
    if len(rgb) == 8:
        if rgb.startswith("00"):
            return None
        rgb = rgb[2:]
    elif len(rgb) != 6:
        return None
    tint = getattr(color_obj, 'tint', 0.0)
    if tint and tint != 0.0:
        try:
            r = int(rgb[0:2], 16)
            g = int(rgb[2:4], 16)
            b = int(rgb[4:6], 16)
            if tint > 0:
                r = int(r + (255 - r) * tint)
                g = int(g + (255 - g) * tint)
                b = int(b + (255 - b) * tint)
            else:
                r = int(r * (1 + tint))
                g = int(g * (1 + tint))
                b = int(b * (1 + tint))
            r = max(0, min(255, r))
            g = max(0, min(255, g))
            b = max(0, min(255, b))
            rgb = f"{r:02X}{g:02X}{b:02X}"
        except Exception:
            pass
    return rgb
